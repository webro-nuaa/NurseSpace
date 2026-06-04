from routes.admin import admin_bp
import re
from datetime import datetime, timezone
from io import BytesIO

from flask import jsonify, request, send_file, current_app
from flask_login import current_user
from utils.auth import login_or_jwt_required
from utils.decorators import admin_required
from models import User, LearningRecord, WrongQuestion, ExamRecord, PointRecord, db
from sqlalchemy import desc, func


@admin_bp.route('/users')
@login_or_jwt_required
@admin_required
def get_users():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    role = request.args.get('role', 'nurse')
    status = request.args.get('status')
    search = request.args.get('search', '').strip()

    query = User.query.filter_by(role=role)

    if status:
        query = query.filter_by(status=status)

    if search:
        query = query.filter(
            db.or_(
                User.username.contains(search),
                User.real_name.contains(search),
                User.department.contains(search)
            )
        )

    pagination = query.order_by(desc(User.created_at))\
        .paginate(page=page, per_page=per_page, error_out=False)

    users_data = []
    for user in pagination.items:
        learning_count = LearningRecord.query.filter_by(user_id=user.id).count()
        wrong_count = WrongQuestion.query.filter_by(user_id=user.id).count()

        users_data.append({
            'id': user.id,
            'username': user.username,
            'real_name': user.real_name,
            'email': user.email,
            'phone': user.phone,
            'department': user.department,
            'school': user.school,
            'serial_number': user.serial_number,
            'status': user.status,
            'points': user.points,
            'consent_accepted': user.consent_accepted,
            'consent_accepted_at': user.consent_accepted_at.isoformat() if user.consent_accepted_at else None,
            'learning_count': learning_count,
            'wrong_count': wrong_count,
            'created_at': user.created_at.isoformat()
        })

    return jsonify({
        'success': True,
        'data': {
            'users': users_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_prev': pagination.has_prev,
                'has_next': pagination.has_next
            }
        }
    })


@admin_bp.route('/users/<int:user_id>', methods=['GET', 'PUT'])
@login_or_jwt_required
@admin_required
def admin_user_detail(user_id: int):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'}), 404

    if request.method == 'GET':
        return jsonify({'success': True, 'data': {
            'id': user.id,
            'username': user.username,
            'real_name': user.real_name,
            'email': user.email,
            'phone': user.phone,
            'department': user.department,
            'school': user.school,
            'serial_number': user.serial_number,
            'role': user.role,
            'status': user.status,
            'points': user.points,
            'created_at': user.created_at.isoformat()
        }})

    data = request.get_json() or {}
    for field in ['real_name', 'email', 'phone', 'department', 'school', 'serial_number', 'role', 'status']:
        if field in data:
            setattr(user, field, (data.get(field) or None))

    if user.email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', user.email):
        return jsonify({'success': False, 'message': '邮箱格式不正确'})
    if user.phone and not re.match(r'^1[3-9]\d{9}$', user.phone):
        return jsonify({'success': False, 'message': '手机号格式不正确'})

    try:
        db.session.commit()
        return jsonify({'success': True, 'message': '用户信息已更新'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"管理员更新用户信息失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '更新失败，请稍后重试'})


@admin_bp.route('/users/<int:user_id>/progress')
@login_or_jwt_required
@admin_required
def get_user_progress(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'}), 404

    from models import Station, Case, CaseCategory, Exam

    category_progress = db.session.query(
        CaseCategory.name,
        func.count(func.distinct(Station.id)).label('total'),
        func.count(func.distinct(LearningRecord.id)).label('completed')
    ).join(Case, CaseCategory.id == Case.category_id)\
     .join(Station, Case.id == Station.case_id)\
     .outerjoin(LearningRecord, (Station.id == LearningRecord.station_id) & (LearningRecord.user_id == user_id))\
     .filter(Case.case_type == 'learning', Station.station_type == 'assessment')\
     .group_by(CaseCategory.id, CaseCategory.name).all()

    recent = db.session.query(LearningRecord, Station, Case)\
        .join(Station, LearningRecord.station_id == Station.id)\
        .join(Case, Station.case_id == Case.id)\
        .filter(LearningRecord.user_id == user_id)\
        .order_by(desc(LearningRecord.completed_at)).limit(20).all()

    exam_records = ExamRecord.query.filter_by(user_id=user_id)\
        .order_by(desc(ExamRecord.submit_time)).all()

    point_recs = PointRecord.query.filter_by(user_id=user_id)\
        .order_by(desc(PointRecord.created_at)).limit(20).all()

    return jsonify({'success': True, 'data': {
        'user': {
            'id': user.id, 'username': user.username, 'real_name': user.real_name,
            'department': user.department, 'points': user.points, 'status': user.status
        },
        'category_progress': [
            {'category': c, 'total': t, 'completed': co or 0,
             'progress': round((co or 0) / t * 100, 1) if t > 0 else 0}
            for c, t, co in category_progress
        ],
        'recent_records': [
            {'id': r.id, 'case_title': c.title, 'station_name': s.name,
             'score': float(r.score) if r.score else 0,
             'completed_at': r.completed_at.isoformat()}
            for r, s, c in recent
        ],
        'exam_records': [
            {'id': e.id, 'exam_id': e.exam_id, 'total_score': float(e.total_score) if e.total_score else 0,
             'status': e.status, 'submit_time': e.submit_time.isoformat() if e.submit_time else None}
            for e in exam_records
        ],
        'point_records': [
            {'id': p.id, 'points': p.points, 'reason': p.reason,
             'created_at': p.created_at.isoformat()}
            for p in point_recs
        ]
    }})


@admin_bp.route('/users/xlsx-template', methods=['GET'])
@login_or_jwt_required
@admin_required
def users_xlsx_template():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'users'
        headers = ['真实姓名', '科室', '学校', '学号', '邮箱', '手机号', '角色', '状态']
        ws.append(headers)
        ws.append(['张三', '内科', '某某护理学院', '2024001', 'zhangsan@example.com', '13800001111', 'nurse', 'active'])
        ws.append(['李四', '教学部', '某某卫生学校', '2024002', 'lisi@example.com', '13900002222', 'nurse', 'active'])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='用户批量导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        current_app.logger.error(f"生成用户导入模板失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '生成模板失败，请稍后重试'})


@admin_bp.route('/users/batch-import-xlsx', methods=['POST'])
@login_or_jwt_required
@admin_required
def users_batch_import_xlsx():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': '只支持 .xlsx 文件'})

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(f.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'message': '空文件'})
        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}
        required = ['真实姓名']
        for r in required:
            if r not in idx:
                return jsonify({'success': False, 'message': f'缺少列：{r}'})

        def get(row, key, default=''):
            i = idx.get(key)
            return (str(row[i]).strip() if i is not None and row[i] is not None else default)

        def _generate_username():
            year = datetime.now(timezone.utc).strftime('%y')
            prefix = f'NS{year}'
            last = User.query.filter(User.username.like(f'{prefix}%')).order_by(User.username.desc()).first()
            if last:
                seq = int(last.username[-3:]) + 1
            else:
                seq = 1
            return f'{prefix}{seq:03d}'

        def _generate_password(username):
            emp_id = username[2:]
            return f'{emp_id}@ns'

        created, skipped, failed = 0, 0, 0
        new_users = []
        for row in rows[1:]:
            if not row:
                continue
            real_name = get(row, '真实姓名')
            if not real_name:
                failed += 1
                continue
            email = get(row, '邮箱')
            phone = get(row, '手机号')
            department = get(row, '科室')
            school = get(row, '学校')
            serial_number = get(row, '学号')
            role = get(row, '角色', 'nurse') or 'nurse'
            status = get(row, '状态', 'active') or 'active'
            if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                failed += 1
                continue
            if phone and not re.match(r'^1[3-9]\d{9}$', phone):
                failed += 1
                continue
            if role not in ['nurse', 'admin']:
                role = 'nurse'
            if status not in ['active', 'disabled']:
                status = 'active'

            try:
                username = _generate_username()
                password = _generate_password(username)
                user = User(username=username, real_name=real_name, email=email or None,
                            phone=phone or None, department=department or None,
                            school=school or None, serial_number=serial_number or None,
                            role=role, status=status)
                user.set_password(password)
                db.session.add(user)
                new_users.append({'username': username, 'password': password, 'real_name': real_name})
                created += 1
            except Exception:
                failed += 1
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'导入完成：新建 {created}，跳过 {skipped}，失败 {failed}',
            'users': new_users
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"用户批量导入失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '导入失败，请稍后重试'})
