from routes.admin import admin_bp
import os
import re
import tempfile
import shutil
import zipfile
import rarfile
import logging
from io import BytesIO
from datetime import datetime, timezone

from flask import jsonify, request, send_file, current_app
from flask_login import current_user
from utils.auth import login_or_jwt_required
from utils.decorators import admin_required
from utils.docx_parser import DocxParser
from models import Case, CaseCategory, Station, StandardAnswer, LearningRecord, ExtensionVideo, ExtensionLink, db
from sqlalchemy import desc, func

docx_parser = DocxParser()


@admin_bp.route('/cases', methods=['GET', 'POST'])
@login_or_jwt_required
@admin_required
def manage_cases():
    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        category_id = request.args.get('category_id', type=int)
        search = request.args.get('search', '').strip()

        case_type_filter = request.args.get('case_type', '').strip()
        include_stations = request.args.get('include_stations', '').strip() == 'true'

        query = db.session.query(Case, CaseCategory)\
            .join(CaseCategory, Case.category_id == CaseCategory.id)

        if category_id:
            query = query.filter(Case.category_id == category_id)

        if case_type_filter in ('learning', 'exam'):
            query = query.filter(Case.case_type == case_type_filter)

        if search:
            query = query.filter(
                db.or_(
                    Case.title.contains(search),
                    CaseCategory.name.contains(search)
                )
            )

        pagination = query.order_by(desc(Case.created_at))\
            .paginate(page=page, per_page=per_page, error_out=False)

        cases_data = []
        for case, category in pagination.items:
            case_stations = Station.query.filter_by(case_id=case.id)\
                .order_by(Station.order_index).all()
            station_count = len(case_stations)
            learning_count = db.session.query(LearningRecord)\
                .join(Station, LearningRecord.station_id == Station.id)\
                .filter(Station.case_id == case.id).count()

            case_item = {
                'id': case.id,
                'title': case.title,
                'category_name': category.name,
                'difficulty': case.difficulty or 'intermediate',
                'case_type': case.case_type or 'learning',
                'station_count': station_count,
                'learning_count': learning_count,
                'created_at': case.created_at.isoformat()
            }
            if include_stations:
                case_item['stations'] = [{
                    'id': s.id,
                    'name': s.name,
                    'question': s.question,
                    'assessment_task': s.assessment_task,
                    'order_index': s.order_index,
                    'standard_answers': [{
                        'answer_item': a.answer_item,
                        'score_weight': float(a.score_weight or 1.0)
                    } for a in s.standard_answers]
                } for s in case_stations]
            cases_data.append(case_item)

        categories = CaseCategory.query.all()
        categories_data = [
            {'id': cat.id, 'name': cat.name, 'description': cat.description}
            for cat in categories
        ]

        return jsonify({
            'success': True,
            'data': {
                'cases': cases_data,
                'categories': categories_data,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': pagination.total,
                    'pages': pagination.pages,
                    'has_prev': pagination.has_prev,
                    'has_next': pagination.has_next
                }
            }
        })

    # POST: JSON body or file upload
    if request.is_json:
        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        category_id = data.get('category_id')
        category_name = (data.get('category_name') or '').strip()
        if not title:
            return jsonify({'success': False, 'message': '标题不能为空'})
        if not category_id and not category_name:
            return jsonify({'success': False, 'message': '请选择类别或输入新类别名称'})
        if category_id:
            category = db.session.get(CaseCategory, category_id)
            if not category:
                return jsonify({'success': False, 'message': '类别不存在'})
        else:
            category = CaseCategory.query.filter_by(name=category_name).first()
            if category:
                category_id = category.id
            else:
                category = CaseCategory(name=category_name)
                db.session.add(category)
                db.session.flush()
                category_id = category.id
        case = Case(
            category_id=category_id, title=title,
            case_guide=(data.get('case_guide') or '').strip(),
            difficulty=data.get('difficulty', 'intermediate'),
            case_type=data.get('case_type', 'learning'),
            file_path=''
        )
        db.session.add(case)
        db.session.flush()

        for si, s_data in enumerate(data.get('stations') or []):
            station = Station(
                case_id=case.id,
                name=(s_data.get('name') or '').strip(),
                assessment_task=(s_data.get('assessment_task') or '').strip(),
                question=(s_data.get('question') or '').strip(),
                order_index=s_data.get('order_index', si)
            )
            db.session.add(station)
            db.session.flush()
            for ai, a_data in enumerate(s_data.get('standard_answers') or []):
                ans = StandardAnswer(
                    station_id=station.id,
                    answer_item=(a_data.get('answer_item') or '').strip(),
                    score_weight=float(a_data.get('score_weight', 1.0)),
                    order_index=ai
                )
                db.session.add(ans)

        for vi, v_data in enumerate(data.get('videos') or []):
            video = ExtensionVideo(
                case_id=case.id,
                title=(v_data.get('title') or '').strip(),
                url=(v_data.get('url') or '').strip(),
                description=(v_data.get('description') or '').strip(),
                order_index=v_data.get('order_index', vi)
            )
            db.session.add(video)

        for li, l_data in enumerate(data.get('links') or []):
            link = ExtensionLink(
                case_id=case.id,
                title=(l_data.get('title') or '').strip(),
                url=(l_data.get('url') or '').strip(),
                description=(l_data.get('description') or '').strip(),
                order_index=l_data.get('order_index', li)
            )
            db.session.add(link)

        for k_data in data.get('extended_knowledge') or []:
            sk = Station(
                case_id=case.id,
                question=(k_data.get('question') or '').strip(),
                station_type='knowledge',
                order_index=0
            )
            db.session.add(sk)
            db.session.flush()
            for idx, a_data in enumerate(k_data.get('answers') or []):
                db.session.add(StandardAnswer(
                    station_id=sk.id,
                    answer_item=(a_data.get('answer_item') or '').strip(),
                    score_weight=float(a_data.get('score_weight', 1)),
                    order_index=idx
                ))

        db.session.commit()
        station_count = Station.query.filter_by(case_id=case.id).count()
        return jsonify({
            'success': True,
            'message': f'案例创建成功（含{station_count}个站点）',
            'case': {'id': case.id, 'title': case.title}
        })

    # File upload
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})

    if not file.filename.endswith('.docx'):
        return jsonify({'success': False, 'message': '只支持docx格式的文件'})

    try:
        cases_dir = current_app.config['CASES_DIR']
        os.makedirs(cases_dir, exist_ok=True)
        orig_name = os.path.basename(file.filename.strip())
        if not orig_name.lower().endswith('.docx'):
            return jsonify({'success': False, 'message': '文件后缀必须为 .docx'})

        target_path = os.path.join(cases_dir, orig_name)
        if os.path.exists(target_path):
            name, ext = os.path.splitext(orig_name)
            idx = 1
            while True:
                candidate = f"{name}_{idx}{ext}"
                target_path = os.path.join(cases_dir, candidate)
                if not os.path.exists(target_path):
                    break
                idx += 1
        file.save(target_path)

        case = docx_parser.parse_file(target_path)

        return jsonify({
            'success': True,
            'message': '案例上传并解析成功',
            'case': {
                'id': case.id,
                'title': case.title,
                'category_name': case.category.name,
                'file_path': target_path
            }
        })

    except Exception as e:
        current_app.logger.error(f"案例上传失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '上传失败，请稍后重试'})


@admin_bp.route('/cases/<int:case_id>', methods=['PUT', 'DELETE'])
@login_or_jwt_required
@admin_required
def update_or_delete_case(case_id: int):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404

    if request.method == 'DELETE':
        try:
            db.session.delete(case)
            db.session.commit()
            return jsonify({'success': True, 'message': '案例已删除'})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"案例删除失败: {e}", exc_info=True)
            return jsonify({'success': False, 'message': '删除失败，请稍后重试'})

    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    case_guide = (data.get('case_guide') or '').strip()
    category_id = data.get('category_id')

    if title:
        case.title = title
    case.case_guide = case_guide
    if category_id:
        category = db.session.get(CaseCategory, category_id)
        if not category:
            return jsonify({'success': False, 'message': '类别不存在'})
        case.category_id = category_id
    if 'difficulty' in data and data['difficulty'] in ('basic', 'intermediate', 'advanced'):
        case.difficulty = data['difficulty']
    if 'case_type' in data and data['case_type'] in ('learning', 'exam'):
        case.case_type = data['case_type']

    try:
        db.session.commit()
        return jsonify({'success': True, 'message': '案例已更新'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"案例更新失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '更新失败，请稍后重试'})


@admin_bp.route('/cases/<int:case_id>')
@login_or_jwt_required
@admin_required
def get_case_detail(case_id):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404

    stations = Station.query.filter_by(case_id=case_id).order_by(Station.order_index).all()
    stations_data = []

    for station in stations:
        answers = StandardAnswer.query.filter_by(station_id=station.id)\
            .order_by(StandardAnswer.order_index).all()

        learning_count = LearningRecord.query.filter_by(station_id=station.id).count()
        avg_score = db.session.query(func.avg(LearningRecord.score))\
            .filter_by(station_id=station.id).scalar()

        stations_data.append({
            'id': station.id,
            'name': station.name or '',
            'assessment_task': station.assessment_task,
            'condition_report': station.condition_report,
            'question': station.question,
            'station_type': station.station_type,
            'answers': [
                {
                    'id': ans.id,
                    'answer_item': ans.answer_item,
                    'score_weight': float(ans.score_weight),
                    'order_index': ans.order_index
                }
                for ans in answers
            ],
            'learning_count': learning_count,
            'avg_score': float(avg_score) if avg_score else 0
        })

    videos = ExtensionVideo.query.filter_by(case_id=case_id).order_by(ExtensionVideo.order_index).all()
    links = ExtensionLink.query.filter_by(case_id=case_id).order_by(ExtensionLink.order_index).all()
    return jsonify({
        'success': True,
        'data': {
            'case': {
                'id': case.id,
                'title': case.title,
                'case_guide': case.case_guide,
                'category_name': case.category.name,
                'difficulty': case.difficulty or 'intermediate',
                'case_type': case.case_type or 'learning',
                'file_path': case.file_path,
                'created_at': case.created_at.isoformat()
            },
            'stations': stations_data,
            'videos': [{'id': v.id, 'title': v.title, 'url': v.url,
                        'description': v.description or '', 'order_index': v.order_index}
                       for v in videos],
            'links': [{'id': l.id, 'title': l.title, 'url': l.url,
                       'description': l.description or '', 'order_index': l.order_index}
                      for l in links]
        }
    })


@admin_bp.route('/cases/batch-delete', methods=['POST'])
@login_or_jwt_required
@admin_required
def batch_delete_cases():
    data = request.get_json() or {}
    ids = data.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': '请提供要删除的案例ID列表'})
    try:
        for cid in ids:
            case = db.session.get(Case, cid)
            if case:
                db.session.delete(case)
        db.session.commit()
        return jsonify({'success': True, 'message': f'已删除 {len(ids)} 个案例'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量删除案例失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '批量删除失败，请稍后重试'})


@admin_bp.route('/cases/batch-upload', methods=['POST'])
@login_or_jwt_required
@admin_required
def batch_upload_cases():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择压缩包文件'})
    f = request.files['file']
    fname_lower = f.filename.lower()
    is_zip = fname_lower.endswith('.zip')
    is_rar = fname_lower.endswith('.rar')
    if not is_zip and not is_rar:
        return jsonify({'success': False, 'message': '只支持 .zip 或 .rar 格式的压缩包'})

    tmp_dir = tempfile.mkdtemp()
    try:
        archive_path = os.path.join(tmp_dir, 'upload.zip' if is_zip else 'upload.rar')
        f.save(archive_path)

        if is_zip:
            try:
                with zipfile.ZipFile(archive_path, 'r') as zf:
                    real_tmp = os.path.realpath(tmp_dir)
                    for member in zf.namelist():
                        member_real = os.path.realpath(os.path.join(tmp_dir, member))
                        if not member_real.startswith(real_tmp + os.sep) and member_real != real_tmp:
                            return jsonify({'success': False, 'message': 'ZIP文件包含非法路径，已拒绝'})
                    zf.extractall(tmp_dir)
            except zipfile.BadZipFile:
                return jsonify({'success': False, 'message': '无效的ZIP文件，请检查压缩包是否完整'})
        else:
            try:
                with rarfile.RarFile(archive_path, 'r') as rf:
                    real_tmp = os.path.realpath(tmp_dir)
                    for member in rf.namelist():
                        member_real = os.path.realpath(os.path.join(tmp_dir, member))
                        if not member_real.startswith(real_tmp + os.sep) and member_real != real_tmp:
                            return jsonify({'success': False, 'message': 'RAR文件包含非法路径，已拒绝'})
                    rf.extractall(tmp_dir)
            except rarfile.BadRarFile:
                return jsonify({'success': False, 'message': '无效的RAR文件，请检查压缩包是否完整'})
            except rarfile.NotRarFile:
                return jsonify({'success': False, 'message': '文件不是有效的RAR格式'})

        os.remove(archive_path)

        results = []
        errors = []
        skipped = []
        total_in_archive = 0
        for root, dirs, files in os.walk(tmp_dir):
            dirs[:] = [d for d in dirs if d != '__MACOSX']
            for fname in files:
                total_in_archive += 1
                if not fname.lower().endswith('.docx'):
                    continue
                if fname.startswith('~') or fname.startswith('._'):
                    skipped.append(fname)
                    continue
                fpath = os.path.join(root, fname)
                try:
                    case = docx_parser.parse_file(fpath)
                    results.append({
                        'filename': fname,
                        'case_id': case.id,
                        'case_title': case.title,
                        'status': 'success'
                    })
                except Exception as e:
                    errors.append({'filename': fname, 'error': str(e), 'status': 'error'})

        total_found = len(results) + len(errors)
        return jsonify({
            'success': True,
            'message': f'批量上传完成，共扫描 {total_found} 个文件，成功: {len(results)}个，失败: {len(errors)}个',
            'data': {
                'total_in_archive': total_in_archive,
                'total_found': total_found,
                'success_count': len(results),
                'error_count': len(errors),
                'results': results,
                'errors': errors
            }
        })

    except Exception as e:
        current_app.logger.error(f"案例批量上传失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '批量上传失败，请稍后重试'})
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@admin_bp.route('/cases/xlsx-template', methods=['GET'])
@login_or_jwt_required
@admin_required
def download_cases_xlsx_template():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'cases'
        headers = ['类别', '案例标题', '站点', '案例指引', '站点名称', '考核任务', '题目',
                   '答案项1', '答案项2', '答案项3', '知识问1', '知识答1', '知识问2', '知识答2']
        ws.append(headers)
        ws.append(['儿科模块', '案例：新生儿黄疸（东22区新生儿科）', '东22区新生儿科',
                   '这里填写案例指引文本', '护理评估', '有条理采集病史；选择性体格评估',
                   '请写出护理评估要点', '评估胎龄与喂养', '评估皮肤黄染范围', '评估家长认知',
                   '病理性黄疸的特点是什么？', '出生24h内出现；程度重；持续时间长', '', ''])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='案例批量导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        current_app.logger.error(f"生成案例导入模板失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '生成模板失败，请稍后重试'})


@admin_bp.route('/cases/batch-import-xlsx', methods=['POST'])
@login_or_jwt_required
@admin_required
def batch_import_cases_xlsx():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': '只支持 .xlsx 文件'})

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(f.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'message': '空文件'})
        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}

        required = ['类别', '案例标题']
        for r in required:
            if r not in idx:
                return jsonify({'success': False, 'message': f'缺少列：{r}'})

        def get(row, key, default=''):
            i = idx.get(key)
            return (str(row[i]).strip() if i is not None and row[i] is not None else default)

        answer_cols = [k for k in header if re.match(r'^答案项\d+$', str(k))]
        q_cols = [k for k in header if re.match(r'^知识问\d+$', str(k))]

        created, skipped = 0, 0
        for row in rows[1:]:
            if not row:
                continue
            category_name = get(row, '类别')
            case_title = get(row, '案例标题')
            if not category_name or not case_title:
                continue

            category = CaseCategory.query.filter_by(name=category_name).first()
            if not category:
                category = CaseCategory(name=category_name, description=f"{category_name}相关医疗案例")
                db.session.add(category)
                db.session.flush()

            existing = Case.query.filter_by(title=case_title, category_id=category.id).first()
            if existing:
                skipped += 1
                continue

            case = Case(category_id=category.id, title=case_title,
                        case_guide=get(row, '案例指引'), file_path='')
            db.session.add(case)
            db.session.flush()

            station_name = get(row, '站点名称')
            if station_name:
                station = Station(case_id=case.id, name=station_name,
                                  assessment_task=get(row, '考核任务'), question=get(row, '题目'))
                db.session.add(station)
                db.session.flush()
                order = 0
                for col in answer_cols:
                    val = get(row, col)
                    if val:
                        db.session.add(StandardAnswer(station_id=station.id, answer_item=val, order_index=order))
                        order += 1

            for qc in q_cols:
                suffix = qc.replace('知识问', '')
                ac = f'知识答{suffix}'
                qv = get(row, qc)
                av = get(row, ac)
                if qv:
                    sk = Station(case_id=case.id, question=qv, station_type='knowledge', order_index=0)
                    db.session.add(sk)
                    db.session.flush()
                    if av:
                        db.session.add(StandardAnswer(station_id=sk.id, answer_item=av, order_index=0))

            created += 1

        db.session.commit()
        return jsonify({'success': True, 'message': f'导入完成：新建 {created}，跳过 {skipped}'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"案例导入失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '导入失败，请稍后重试'})
