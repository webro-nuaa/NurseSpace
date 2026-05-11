from flask import Blueprint, request, jsonify, render_template, send_file, current_app
from flask_login import current_user
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import csrf
from utils.auth import login_or_jwt_required
from utils.decorators import admin_required
from models import User, Case, CaseCategory, Station, StandardAnswer, LearningRecord, WrongQuestion, Exam, ExamQuestion, ExamAnswer, ExamRecord, PointRecord, ExtendedKnowledge, KnowledgeAnswer, ExtensionVideo, ExtensionLink, AiSetting, BaiduAsrKey, db
from utils.docx_parser import DocxParser
from utils.crypto import encrypt_value, decrypt_value
from sqlalchemy import desc, func
from datetime import datetime, timezone
import os
import tempfile
import shutil
import zipfile
import logging
import rarfile
from io import BytesIO
import re

admin_bp = Blueprint('admin', __name__)
docx_parser = DocxParser()


@admin_bp.route('/dashboard')
@login_or_jwt_required
@admin_required
def dashboard():
    total_users = User.query.filter_by(role='nurse').count()
    active_users = User.query.filter_by(role='nurse', status='active').count()
    total_cases = Case.query.count()
    total_stations = Station.query.count()
    total_learning_records = LearningRecord.query.count()
    total_exams = Exam.query.count()

    recent_activities = db.session.query(LearningRecord, User, Station, Case)\
        .join(User, LearningRecord.user_id == User.id)\
        .join(Station, LearningRecord.station_id == Station.id)\
        .join(Case, Station.case_id == Case.id)\
        .order_by(desc(LearningRecord.completed_at))\
        .limit(10).all()

    activities_data = []
    for record, user, station, case in recent_activities:
        activities_data.append({
            'id': record.id,
            'user_name': user.real_name,
            'case_title': case.title,
            'station_name': station.name,
            'score': float(record.score) if record.score else 0,
            'completed_at': record.completed_at.isoformat()
        })

    category_stats = db.session.query(
        CaseCategory.name,
        func.count(func.distinct(Case.id)).label('case_count'),
        func.count(func.distinct(Station.id)).label('station_count')
    ).join(Case, CaseCategory.id == Case.category_id)\
     .outerjoin(Station, Case.id == Station.case_id)\
     .group_by(CaseCategory.id, CaseCategory.name).all()

    category_data = []
    for category_name, case_count, station_count in category_stats:
        category_data.append({
            'category': category_name,
            'case_count': case_count,
            'station_count': station_count
        })

    learning_stats = db.session.query(
        func.avg(LearningRecord.score).label('avg_score'),
        func.count(LearningRecord.id).label('total_records'),
        func.count(WrongQuestion.id).label('wrong_count')
    ).outerjoin(WrongQuestion, LearningRecord.station_id == WrongQuestion.station_id).first()

    return jsonify({
        'success': True,
        'data': {
            'statistics': {
                'total_users': total_users,
                'active_users': active_users,
                'total_cases': total_cases,
                'total_stations': total_stations,
                'total_learning_records': total_learning_records,
                'total_exams': total_exams,
                'avg_score': float(learning_stats.avg_score) if learning_stats.avg_score else 0,
                'wrong_rate': round(learning_stats.wrong_count / learning_stats.total_records * 100, 1) if learning_stats.total_records > 0 else 0
            },
            'recent_activities': activities_data,
            'category_data': category_data
        }
    })


@admin_bp.route('/ai-settings', methods=['GET', 'PUT'])
@login_or_jwt_required
@admin_required
def ai_settings():
    setting = AiSetting.get_singleton()

    if request.method == 'GET':
        def _mask_key(key):
            if not key:
                return ''
            try:
                plain = decrypt_value(key)
            except Exception:
                return '***'
            if not plain:
                return ''
            return '***' + plain[-4:] if len(plain) > 4 else '***'
        return jsonify({'success': True, 'data': {
            'provider': setting.provider,
            'openai_key': _mask_key(setting.openai_key),
            'openai_model': setting.openai_model or '',
            'openai_base_url': setting.openai_base_url or '',
            'zhipu_key': _mask_key(setting.zhipu_key),
            'zhipu_model': setting.zhipu_model or '',
            'zhipu_base_url': setting.zhipu_base_url or ''
        }})

    data = request.get_json() or {}
    provider = data.get('provider') or setting.provider
    if provider not in ['glm', 'openai', 'local']:
        return jsonify({'success': False, 'message': 'provider 取值必须是 glm | openai | local'})

    setting.provider = provider
    for field in ['openai_key', 'openai_model', 'openai_base_url', 'zhipu_key', 'zhipu_model', 'zhipu_base_url']:
        if field in data:
            value = data.get(field) or None
            # 加密 key 字段，其他字段不加密
            if field in ('openai_key', 'zhipu_key') and value:
                if value.startswith('***'):
                    continue
                value = encrypt_value(value)
            setattr(setting, field, value)
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': 'AI设置已更新'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"AI设置保存失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '保存失败，请稍后重试'})


@admin_bp.route('/users')
@login_or_jwt_required
@admin_required
def get_users():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    role = request.args.get('role', 'nurse')
    status = request.args.get('status')
    search = request.args.get('search', '').strip()

    query = User.query.filter_by(role=role)

    if status:
        query = query.filter_by(status=status)

    if search:
        query = query.filter(
            db.or_(
                User.username.contains(search),
                User.real_name.contains(search),
                User.department.contains(search)
            )
        )

    pagination = query.order_by(desc(User.created_at))\
        .paginate(page=page, per_page=per_page, error_out=False)

    users_data = []
    for user in pagination.items:
        learning_count = LearningRecord.query.filter_by(user_id=user.id).count()
        wrong_count = WrongQuestion.query.filter_by(user_id=user.id).count()

        users_data.append({
            'id': user.id,
            'username': user.username,
            'real_name': user.real_name,
            'email': user.email,
            'phone': user.phone,
            'department': user.department,
            'status': user.status,
            'points': user.points,
            'learning_count': learning_count,
            'wrong_count': wrong_count,
            'created_at': user.created_at.isoformat()
        })

    return jsonify({
        'success': True,
        'data': {
            'users': users_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_prev': pagination.has_prev,
                'has_next': pagination.has_next
            }
        }
    })


@admin_bp.route('/users/<int:user_id>', methods=['GET', 'PUT'])
@login_or_jwt_required
@admin_required
def admin_user_detail(user_id: int):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'}), 404

    if request.method == 'GET':
        return jsonify({'success': True, 'data': {
            'id': user.id,
            'username': user.username,
            'real_name': user.real_name,
            'email': user.email,
            'phone': user.phone,
            'department': user.department,
            'role': user.role,
            'status': user.status
        }})

    data = request.get_json() or {}
    for field in ['real_name', 'email', 'phone', 'department', 'role', 'status']:
        if field in data:
            setattr(user, field, (data.get(field) or None))

    if user.email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', user.email):
        return jsonify({'success': False, 'message': '邮箱格式不正确'})
    if user.phone and not re.match(r'^1[3-9]\d{9}$', user.phone):
        return jsonify({'success': False, 'message': '手机号格式不正确'})

    try:
        db.session.commit()
        return jsonify({'success': True, 'message': '用户信息已更新'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"管理员更新用户信息失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '更新失败，请稍后重试'})


@admin_bp.route('/users/xlsx-template', methods=['GET'])
@login_or_jwt_required
@admin_required
def users_xlsx_template():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'users'
        headers = ['真实姓名', '科室', '邮箱', '手机号', '角色', '状态']
        ws.append(headers)
        ws.append(['张三', '内科', 'zhangsan@example.com', '13800001111', 'nurse', 'active'])
        ws.append(['李四', '教学部', 'lisi@example.com', '13900002222', 'nurse', 'active'])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='用户批量导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        current_app.logger.error(f"生成用户导入模板失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '生成模板失败，请稍后重试'})


@admin_bp.route('/users/batch-import-xlsx', methods=['POST'])
@login_or_jwt_required
@admin_required
@csrf.exempt
def users_batch_import_xlsx():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': '只支持 .xlsx 文件'})

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(f.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'message': '空文件'})
        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}
        required = ['真实姓名']
        for r in required:
            if r not in idx:
                return jsonify({'success': False, 'message': f'缺少列：{r}'})

        def get(row, key, default=''):
            i = idx.get(key)
            return (str(row[i]).strip() if i is not None and row[i] is not None else default)

        def _generate_username():
            year = datetime.now(timezone.utc).strftime('%y')
            prefix = f'NS{year}'
            last = User.query.filter(User.username.like(f'{prefix}%')).order_by(User.username.desc()).first()
            if last:
                seq = int(last.username[-3:]) + 1
            else:
                seq = 1
            return f'{prefix}{seq:03d}'

        def _generate_password(username):
            emp_id = username[2:]
            return f'{emp_id}@ns'

        created, skipped, failed = 0, 0, 0
        new_users = []
        for row in rows[1:]:
            if not row:
                continue
            real_name = get(row, '真实姓名')
            if not real_name:
                failed += 1
                continue
            email = get(row, '邮箱')
            phone = get(row, '手机号')
            department = get(row, '科室')
            role = get(row, '角色', 'nurse') or 'nurse'
            status = get(row, '状态', 'active') or 'active'
            if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                failed += 1
                continue
            if phone and not re.match(r'^1[3-9]\d{9}$', phone):
                failed += 1
                continue
            if role not in ['nurse', 'admin']:
                role = 'nurse'
            if status not in ['active', 'disabled']:
                status = 'active'

            try:
                username = _generate_username()
                password = _generate_password(username)
                user = User(username=username, real_name=real_name, email=email or None,
                            phone=phone or None, department=department or None, role=role, status=status)
                user.set_password(password)
                db.session.add(user)
                new_users.append({'username': username, 'password': password, 'real_name': real_name})
                created += 1
            except Exception:
                failed += 1
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'导入完成：新建 {created}，跳过 {skipped}，失败 {failed}',
            'users': new_users
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"用户批量导入失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '导入失败，请稍后重试'})


@admin_bp.route('/cases', methods=['GET', 'POST'])
@login_or_jwt_required
@admin_required
@csrf.exempt  # POST is file upload (multipart/form-data)
def manage_cases():
    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        category_id = request.args.get('category_id', type=int)
        search = request.args.get('search', '').strip()

        case_type_filter = request.args.get('case_type', '').strip()
        include_stations = request.args.get('include_stations', '').strip() == 'true'

        query = db.session.query(Case, CaseCategory)\
            .join(CaseCategory, Case.category_id == CaseCategory.id)

        if category_id:
            query = query.filter(Case.category_id == category_id)

        if case_type_filter in ('learning', 'exam'):
            query = query.filter(Case.case_type == case_type_filter)

        if search:
            query = query.filter(
                db.or_(
                    Case.title.contains(search),
                    CaseCategory.name.contains(search)
                )
            )

        pagination = query.order_by(desc(Case.created_at))\
            .paginate(page=page, per_page=per_page, error_out=False)

        cases_data = []
        for case, category in pagination.items:
            case_stations = Station.query.filter_by(case_id=case.id)\
                .order_by(Station.order_index).all()
            station_count = len(case_stations)
            learning_count = db.session.query(LearningRecord)\
                .join(Station, LearningRecord.station_id == Station.id)\
                .filter(Station.case_id == case.id).count()

            case_item = {
                'id': case.id,
                'title': case.title,
                'category_name': category.name,
                'difficulty': case.difficulty or 'intermediate',
                'case_type': case.case_type or 'learning',
                'station_count': station_count,
                'learning_count': learning_count,
                'created_at': case.created_at.isoformat()
            }
            if include_stations:
                case_item['stations'] = [{
                    'id': s.id,
                    'name': s.name,
                    'question': s.question,
                    'assessment_task': s.assessment_task,
                    'order_index': s.order_index,
                    'standard_answers': [{
                        'answer_item': a.answer_item,
                        'score_weight': float(a.score_weight or 1.0)
                    } for a in s.standard_answers]
                } for s in case_stations]
            cases_data.append(case_item)

        categories = CaseCategory.query.all()
        categories_data = [
            {'id': cat.id, 'name': cat.name, 'description': cat.description}
            for cat in categories
        ]

        return jsonify({
            'success': True,
            'data': {
                'cases': cases_data,
                'categories': categories_data,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': pagination.total,
                    'pages': pagination.pages,
                    'has_prev': pagination.has_prev,
                    'has_next': pagination.has_next
                }
            }
        })

    # POST: 上传案例文档 或 JSON 手动创建（支持完整案例含子元素）
    # JSON body 手动创建（可包含 stations/videos/links/extended_knowledge）
    if request.is_json:
        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        category_id = data.get('category_id')
        category_name = (data.get('category_name') or '').strip()
        if not title:
            return jsonify({'success': False, 'message': '标题不能为空'})
        if not category_id and not category_name:
            return jsonify({'success': False, 'message': '请选择类别或输入新类别名称'})
        if category_id:
            category = db.session.get(CaseCategory, category_id)
            if not category:
                return jsonify({'success': False, 'message': '类别不存在'})
        else:
            category = CaseCategory.query.filter_by(name=category_name).first()
            if category:
                category_id = category.id
            else:
                category = CaseCategory(name=category_name)
                db.session.add(category)
                db.session.flush()
                category_id = category.id
        case = Case(
            category_id=category_id, title=title,
            case_guide=(data.get('case_guide') or '').strip(),
            difficulty=data.get('difficulty', 'intermediate'),
            case_type=data.get('case_type', 'learning'),
            file_path=''
        )
        db.session.add(case)
        db.session.flush()

        # 创建站点（含标准答案）
        for si, s_data in enumerate(data.get('stations') or []):
            station = Station(
                case_id=case.id,
                name=(s_data.get('name') or '').strip(),
                assessment_task=(s_data.get('assessment_task') or '').strip(),
                question=(s_data.get('question') or '').strip(),
                order_index=s_data.get('order_index', si)
            )
            db.session.add(station)
            db.session.flush()
            for ai, a_data in enumerate(s_data.get('standard_answers') or []):
                ans = StandardAnswer(
                    station_id=station.id,
                    answer_item=(a_data.get('answer_item') or '').strip(),
                    score_weight=float(a_data.get('score_weight', 1.0)),
                    order_index=ai
                )
                db.session.add(ans)

        # 创建扩展视频
        for vi, v_data in enumerate(data.get('videos') or []):
            video = ExtensionVideo(
                case_id=case.id,
                title=(v_data.get('title') or '').strip(),
                url=(v_data.get('url') or '').strip(),
                description=(v_data.get('description') or '').strip(),
                order_index=v_data.get('order_index', vi)
            )
            db.session.add(video)

        # 创建扩展链接
        for li, l_data in enumerate(data.get('links') or []):
            link = ExtensionLink(
                case_id=case.id,
                title=(l_data.get('title') or '').strip(),
                url=(l_data.get('url') or '').strip(),
                description=(l_data.get('description') or '').strip(),
                order_index=l_data.get('order_index', li)
            )
            db.session.add(link)

        # 创建扩展知识
        for k_data in data.get('extended_knowledge') or []:
            ek = ExtendedKnowledge(
                case_id=case.id,
                question=(k_data.get('question') or '').strip()
            )
            db.session.add(ek)
            db.session.flush()  # 获取 ek.id
            for idx, a_data in enumerate(k_data.get('answers') or []):
                db.session.add(KnowledgeAnswer(
                    knowledge_id=ek.id,
                    answer_item=(a_data.get('answer_item') or '').strip(),
                    score_weight=float(a_data.get('score_weight', 1)),
                    order_index=idx
                ))

        db.session.commit()
        station_count = Station.query.filter_by(case_id=case.id).count()
        return jsonify({
            'success': True,
            'message': f'案例创建成功（含{station_count}个站点）',
            'case': {'id': case.id, 'title': case.title}
        })

    # 文件上传
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})

    if not file.filename.endswith('.docx'):
        return jsonify({'success': False, 'message': '只支持docx格式的文件'})

    try:
        cases_dir = current_app.config['CASES_DIR']
        os.makedirs(cases_dir, exist_ok=True)
        orig_name = os.path.basename(file.filename.strip())
        if not orig_name.lower().endswith('.docx'):
            return jsonify({'success': False, 'message': '文件后缀必须为 .docx'})

        target_path = os.path.join(cases_dir, orig_name)
        if os.path.exists(target_path):
            name, ext = os.path.splitext(orig_name)
            idx = 1
            while True:
                candidate = f"{name}_{idx}{ext}"
                target_path = os.path.join(cases_dir, candidate)
                if not os.path.exists(target_path):
                    break
                idx += 1
        file.save(target_path)

        case = docx_parser.parse_file(target_path)

        return jsonify({
            'success': True,
            'message': '案例上传并解析成功',
            'case': {
                'id': case.id,
                'title': case.title,
                'category_name': case.category.name,
                'file_path': target_path
            }
        })

    except Exception as e:
        current_app.logger.error(f"案例上传失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '上传失败，请稍后重试'})


@admin_bp.route('/cases/<int:case_id>', methods=['PUT', 'DELETE'])
@login_or_jwt_required
@admin_required
def update_or_delete_case(case_id: int):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404

    if request.method == 'DELETE':
        try:
            db.session.delete(case)
            db.session.commit()
            return jsonify({'success': True, 'message': '案例已删除'})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"案例删除失败: {e}", exc_info=True)
            return jsonify({'success': False, 'message': '删除失败，请稍后重试'})

    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    case_guide = (data.get('case_guide') or '').strip()
    category_id = data.get('category_id')

    if title:
        case.title = title
    case.case_guide = case_guide
    if category_id:
        category = db.session.get(CaseCategory, category_id)
        if not category:
            return jsonify({'success': False, 'message': '类别不存在'})
        case.category_id = category_id
    if 'difficulty' in data and data['difficulty'] in ('basic', 'intermediate', 'advanced'):
        case.difficulty = data['difficulty']
    if 'case_type' in data and data['case_type'] in ('learning', 'exam'):
        case.case_type = data['case_type']

    try:
        db.session.commit()
        return jsonify({'success': True, 'message': '案例已更新'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"案例更新失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '更新失败，请稍后重试'})


@admin_bp.route('/cases/batch-delete', methods=['POST'])
@login_or_jwt_required
@admin_required
def batch_delete_cases():
    data = request.get_json() or {}
    ids = data.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': '请提供要删除的案例ID列表'})
    try:
        for cid in ids:
            case = db.session.get(Case, cid)
            if case:
                db.session.delete(case)
        db.session.commit()
        return jsonify({'success': True, 'message': f'已删除 {len(ids)} 个案例'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量删除案例失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '批量删除失败，请稍后重试'})


@admin_bp.route('/cases/batch-upload', methods=['POST'])
@login_or_jwt_required
@admin_required
@csrf.exempt
def batch_upload_cases():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择压缩包文件'})
    f = request.files['file']
    fname_lower = f.filename.lower()
    is_zip = fname_lower.endswith('.zip')
    is_rar = fname_lower.endswith('.rar')
    if not is_zip and not is_rar:
        return jsonify({'success': False, 'message': '只支持 .zip 或 .rar 格式的压缩包'})

    tmp_dir = tempfile.mkdtemp()
    try:
        archive_path = os.path.join(tmp_dir, 'upload.zip' if is_zip else 'upload.rar')
        f.save(archive_path)

        if is_zip:
            try:
                with zipfile.ZipFile(archive_path, 'r') as zf:
                    real_tmp = os.path.realpath(tmp_dir)
                    for member in zf.namelist():
                        member_real = os.path.realpath(os.path.join(tmp_dir, member))
                        if not member_real.startswith(real_tmp + os.sep) and member_real != real_tmp:
                            return jsonify({'success': False, 'message': 'ZIP文件包含非法路径，已拒绝'})
                    zf.extractall(tmp_dir)
            except zipfile.BadZipFile:
                return jsonify({'success': False, 'message': '无效的ZIP文件，请检查压缩包是否完整'})
        else:
            try:
                with rarfile.RarFile(archive_path, 'r') as rf:
                    real_tmp = os.path.realpath(tmp_dir)
                    for member in rf.namelist():
                        member_real = os.path.realpath(os.path.join(tmp_dir, member))
                        if not member_real.startswith(real_tmp + os.sep) and member_real != real_tmp:
                            return jsonify({'success': False, 'message': 'RAR文件包含非法路径，已拒绝'})
                    rf.extractall(tmp_dir)
            except rarfile.BadRarFile:
                return jsonify({'success': False, 'message': '无效的RAR文件，请检查压缩包是否完整'})
            except rarfile.NotRarFile:
                return jsonify({'success': False, 'message': '文件不是有效的RAR格式'})

        os.remove(archive_path)

        results = []
        errors = []
        skipped = []
        total_in_archive = 0
        for root, dirs, files in os.walk(tmp_dir):
            dirs[:] = [d for d in dirs if d != '__MACOSX']
            for fname in files:
                total_in_archive += 1
                if not fname.lower().endswith('.docx'):
                    continue
                if fname.startswith('~') or fname.startswith('._'):
                    skipped.append(fname)
                    continue
                fpath = os.path.join(root, fname)
                try:
                    case = docx_parser.parse_file(fpath)
                    results.append({
                        'filename': fname,
                        'case_id': case.id,
                        'case_title': case.title,
                        'status': 'success'
                    })
                except Exception as e:
                    errors.append({'filename': fname, 'error': str(e), 'status': 'error'})

        total_found = len(results) + len(errors)
        return jsonify({
            'success': True,
            'message': f'批量上传完成，共扫描 {total_found} 个文件，成功: {len(results)}个，失败: {len(errors)}个',
            'data': {
                'total_in_archive': total_in_archive,
                'total_found': total_found,
                'success_count': len(results),
                'error_count': len(errors),
                'results': results,
                'errors': errors
            }
        })

    except Exception as e:
        current_app.logger.error(f"案例批量上传失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '批量上传失败，请稍后重试'})
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@admin_bp.route('/cases/xlsx-template', methods=['GET'])
@login_or_jwt_required
@admin_required
def download_cases_xlsx_template():
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'cases'
        headers = ['类别', '案例标题', '站点', '案例指引', '站点名称', '考核任务', '题目',
                   '答案项1', '答案项2', '答案项3', '知识问1', '知识答1', '知识问2', '知识答2']
        ws.append(headers)
        ws.append(['儿科模块', '案例：新生儿黄疸（东22区新生儿科）', '东22区新生儿科',
                   '这里填写案例指引文本', '护理评估', '有条理采集病史；选择性体格评估',
                   '请写出护理评估要点', '评估胎龄与喂养', '评估皮肤黄染范围', '评估家长认知',
                   '病理性黄疸的特点是什么？', '出生24h内出现；程度重；持续时间长', '', ''])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='案例批量导入模板.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        current_app.logger.error(f"生成案例导入模板失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '生成模板失败，请稍后重试'})


@admin_bp.route('/cases/batch-import-xlsx', methods=['POST'])
@login_or_jwt_required
@admin_required
@csrf.exempt
def batch_import_cases_xlsx():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': '只支持 .xlsx 文件'})

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(f.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'message': '空文件'})
        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}

        required = ['类别', '案例标题']
        for r in required:
            if r not in idx:
                return jsonify({'success': False, 'message': f'缺少列：{r}'})

        def get(row, key, default=''):
            i = idx.get(key)
            return (str(row[i]).strip() if i is not None and row[i] is not None else default)

        answer_cols = [k for k in header if re.match(r'^答案项\d+$', str(k))]
        q_cols = [k for k in header if re.match(r'^知识问\d+$', str(k))]

        created, skipped = 0, 0
        for row in rows[1:]:
            if not row:
                continue
            category_name = get(row, '类别')
            case_title = get(row, '案例标题')
            if not category_name or not case_title:
                continue

            category = CaseCategory.query.filter_by(name=category_name).first()
            if not category:
                category = CaseCategory(name=category_name, description=f"{category_name}相关医疗案例")
                db.session.add(category)
                db.session.flush()

            existing = Case.query.filter_by(title=case_title, category_id=category.id).first()
            if existing:
                skipped += 1
                continue

            case = Case(category_id=category.id, title=case_title,
                        case_guide=get(row, '案例指引'), file_path='')
            db.session.add(case)
            db.session.flush()

            station_name = get(row, '站点名称')
            if station_name:
                station = Station(case_id=case.id, name=station_name,
                                  assessment_task=get(row, '考核任务'), question=get(row, '题目'))
                db.session.add(station)
                db.session.flush()
                order = 0
                for col in answer_cols:
                    val = get(row, col)
                    if val:
                        db.session.add(StandardAnswer(station_id=station.id, answer_item=val, order_index=order))
                        order += 1

            for qc in q_cols:
                suffix = qc.replace('知识问', '')
                ac = f'知识答{suffix}'
                qv = get(row, qc)
                av = get(row, ac)
                if qv:
                    ek = ExtendedKnowledge(case_id=case.id, question=qv)
                    db.session.add(ek)
                    db.session.flush()
                    if av:
                        db.session.add(KnowledgeAnswer(knowledge_id=ek.id, answer_item=av, order_index=0))

            created += 1

        db.session.commit()
        return jsonify({'success': True, 'message': f'导入完成：新建 {created}，跳过 {skipped}'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"案例导入失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '导入失败，请稍后重试'})


@admin_bp.route('/cases/<int:case_id>')
@login_or_jwt_required
@admin_required
def get_case_detail(case_id):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404

    stations = Station.query.filter_by(case_id=case_id).order_by(Station.order_index).all()
    stations_data = []

    for station in stations:
        answers = StandardAnswer.query.filter_by(station_id=station.id)\
            .order_by(StandardAnswer.order_index).all()

        learning_count = LearningRecord.query.filter_by(station_id=station.id).count()
        avg_score = db.session.query(func.avg(LearningRecord.score))\
            .filter_by(station_id=station.id).scalar()

        stations_data.append({
            'id': station.id,
            'name': station.name,
            'assessment_task': station.assessment_task,
            'question': station.question,
            'answers': [
                {
                    'id': ans.id,
                    'answer_item': ans.answer_item,
                    'score_weight': float(ans.score_weight),
                    'order_index': ans.order_index
                }
                for ans in answers
            ],
            'learning_count': learning_count,
            'avg_score': float(avg_score) if avg_score else 0
        })

    videos = ExtensionVideo.query.filter_by(case_id=case_id).order_by(ExtensionVideo.order_index).all()
    links = ExtensionLink.query.filter_by(case_id=case_id).order_by(ExtensionLink.order_index).all()
    extended_knowledge = ExtendedKnowledge.query.filter_by(case_id=case_id).all()
    knowledge_data = []
    for ek in extended_knowledge:
        answers = KnowledgeAnswer.query.filter_by(knowledge_id=ek.id)\
            .order_by(KnowledgeAnswer.order_index).all()
        knowledge_data.append({
            'id': ek.id,
            'question': ek.question,
            'answers': [{'id': a.id, 'answer_item': a.answer_item,
                         'score_weight': float(a.score_weight)} for a in answers]
        })

    return jsonify({
        'success': True,
        'data': {
            'case': {
                'id': case.id,
                'title': case.title,
                'case_guide': case.case_guide,
                'category_name': case.category.name,
                'difficulty': case.difficulty or 'intermediate',
                'case_type': case.case_type or 'learning',
                'file_path': case.file_path,
                'created_at': case.created_at.isoformat()
            },
            'stations': stations_data,
            'extended_knowledge': knowledge_data,
            'videos': [{'id': v.id, 'title': v.title, 'url': v.url,
                        'description': v.description or '', 'order_index': v.order_index}
                       for v in videos],
            'links': [{'id': l.id, 'title': l.title, 'url': l.url,
                       'description': l.description or '', 'order_index': l.order_index}
                      for l in links]
        }
    })


@admin_bp.route('/exams', methods=['GET', 'POST'])
@login_or_jwt_required
@admin_required
def manage_exams():
    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        pagination = Exam.query.order_by(desc(Exam.created_at))\
            .paginate(page=page, per_page=per_page, error_out=False)

        exams_data = []
        for exam in pagination.items:
            question_count = ExamQuestion.query.filter_by(exam_id=exam.id).count()
            participant_count = ExamRecord.query.filter_by(exam_id=exam.id).count()

            exams_data.append({
                'id': exam.id,
                'title': exam.title,
                'description': exam.description,
                'duration': exam.duration,
                'status': exam.status,
                'question_count': question_count,
                'participant_count': participant_count,
                'start_time': exam.start_time.isoformat() if exam.start_time else None,
                'end_time': exam.end_time.isoformat() if exam.end_time else None,
                'created_at': exam.created_at.isoformat()
            })

        return jsonify({
            'success': True,
            'data': {
                'exams': exams_data,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': pagination.total,
                    'pages': pagination.pages,
                    'has_prev': pagination.has_prev,
                    'has_next': pagination.has_next
                }
            }
        })

    # POST: 创建考试
    data = request.get_json()

    title = data.get('title', '').strip()
    description = data.get('description', '').strip()
    duration = data.get('duration', 60)
    start_time = data.get('start_time')
    end_time = data.get('end_time')

    if not title:
        return jsonify({'success': False, 'message': '考试标题不能为空'})

    try:
        exam = Exam(
            title=title,
            description=description,
            creator_id=current_user.id,
            duration=duration,
            start_time=datetime.fromisoformat(start_time) if start_time else None,
            end_time=datetime.fromisoformat(end_time) if end_time else None
        )

        db.session.add(exam)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '考试创建成功',
            'exam': {'id': exam.id, 'title': exam.title}
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"考试创建失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '创建失败，请稍后重试'})


@admin_bp.route('/exams/<int:exam_id>/questions', methods=['GET', 'POST', 'DELETE'])
@login_or_jwt_required
@admin_required
def manage_exam_questions(exam_id):
    exam = Exam.query.get_or_404(exam_id)

    if request.method == 'DELETE':
        data = request.get_json()
        case_ids = data.get('case_ids', [])
        if not case_ids:
            return jsonify({'success': False, 'message': '请指定要移除的题目'})
        try:
            ExamQuestion.query.filter(
                ExamQuestion.exam_id == exam_id,
                ExamQuestion.case_id.in_(case_ids)
            ).delete(synchronize_session='fetch')
            db.session.commit()
            return jsonify({'success': True, 'message': '已移除'})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"从考试移除案例失败: {e}", exc_info=True)
            return jsonify({'success': False, 'message': '移除失败，请稍后重试'})

    if request.method == 'GET':
        questions = db.session.query(ExamQuestion, Case)\
            .join(Case, ExamQuestion.case_id == Case.id)\
            .filter(ExamQuestion.exam_id == exam_id)\
            .order_by(ExamQuestion.order_index).all()

        questions_data = []
        for eq, case in questions:
            station_count = Station.query.filter_by(case_id=case.id).count()
            questions_data.append({
                'id': eq.id,
                'case_id': case.id,
                'case_title': case.title,
                'difficulty': case.difficulty,
                'score': float(eq.score),
                'order_index': eq.order_index,
                'station_count': station_count
            })

        return jsonify({
            'success': True,
            'data': {
                'exam': {'id': exam.id, 'title': exam.title, 'status': exam.status},
                'questions': questions_data
            }
        })

    # POST: 添加题目
    data = request.get_json()
    case_ids = data.get('case_ids', [])

    if not case_ids:
        return jsonify({'success': False, 'message': '请选择至少一个案例'})

    try:
        max_order = db.session.query(func.max(ExamQuestion.order_index))\
            .filter_by(exam_id=exam_id).scalar() or 0

        for i, case_id in enumerate(case_ids):
            existing = ExamQuestion.query.filter_by(
                exam_id=exam_id,
                case_id=case_id
            ).first()

            if not existing:
                exam_question = ExamQuestion(
                    exam_id=exam_id,
                    case_id=case_id,
                    score=100.0,
                    order_index=max_order + i + 1
                )
                db.session.add(exam_question)

        db.session.commit()

        return jsonify({'success': True, 'message': '案例添加成功'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"考试添加案例失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '添加失败，请稍后重试'})

@admin_bp.route('/exams/<int:exam_id>/questions/clear', methods=['POST'])
@login_or_jwt_required
@admin_required
def clear_exam_questions(exam_id):
    Exam.query.get_or_404(exam_id)
    ExamQuestion.query.filter_by(exam_id=exam_id).delete()
    db.session.commit()
    return jsonify({'success': True, 'message': '已清空所有题目'})


@admin_bp.route('/exams/<int:exam_id>/review', methods=['GET'])
@login_or_jwt_required
@admin_required
def get_exam_review(exam_id):
    """List all participants and their answers for an exam."""
    exam = Exam.query.get_or_404(exam_id)

    records = ExamRecord.query.filter_by(
        exam_id=exam_id, status='submitted'
    ).order_by(ExamRecord.submit_time.desc()).all()

    participants = []
    for record in records:
        user = db.session.get(User, record.user_id)
        answers = ExamAnswer.query.filter_by(
            exam_record_id=record.id
        ).order_by(ExamAnswer.id).all()

        answers_data = []
        for ans in answers:
            station = db.session.get(Station, ans.station_id) if ans.station_id else None
            exam_question = db.session.get(ExamQuestion, ans.exam_question_id) if ans.exam_question_id else None
            case = db.session.get(Case, exam_question.case_id) if exam_question else None

            standard_answers_data = []
            if station:
                standard_answers_data = [
                    {'answer_item': sa.answer_item, 'score_weight': float(sa.score_weight)}
                    for sa in station.standard_answers.all()
                ]

            answers_data.append({
                'id': ans.id,
                'exam_question_id': ans.exam_question_id,
                'station_id': ans.station_id,
                'station_name': station.name if station else '',
                'question': station.question if station else '',
                'user_answer': ans.user_answer,
                'score': float(ans.score) if ans.score else 0,
                'ai_feedback': ans.ai_feedback,
                'standard_answers': standard_answers_data,
                'case_title': case.title if case else '',
                'case_id': case.id if case else None
            })

        participants.append({
            'record_id': record.id,
            'user_id': user.id if user else record.user_id,
            'real_name': user.real_name if user else '未知',
            'username': user.username if user else '',
            'department': user.department if user else '',
            'total_score': float(record.total_score) if record.total_score else 0,
            'max_score': float(record.max_score) if record.max_score else 0,
            'start_time': record.start_time.isoformat() if record.start_time else None,
            'submit_time': record.submit_time.isoformat() if record.submit_time else None,
            'answers': answers_data
        })

    return jsonify({
        'success': True,
        'data': {
            'exam': {'id': exam.id, 'title': exam.title, 'status': exam.status},
            'participants': participants
        }
    })


@admin_bp.route('/exams/<int:exam_id>/review/<int:answer_id>/score', methods=['PUT'])
@login_or_jwt_required
@admin_required
def update_exam_answer_score(exam_id, answer_id):
    """Admin manually adjusts an answer's score and recalculates record total."""
    data = request.get_json() or {}
    new_score = data.get('score')

    if new_score is None:
        return jsonify({'success': False, 'message': '请提供分数'}), 400

    try:
        new_score = float(new_score)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': '分数格式无效'}), 400

    answer = ExamAnswer.query.get_or_404(answer_id)
    answer.score = new_score

    # Recalculate the exam record's total score
    record = db.session.get(ExamRecord, answer.exam_record_id)
    if record:
        all_answers = ExamAnswer.query.filter_by(exam_record_id=record.id).all()
        record.total_score = sum(float(a.score or 0) for a in all_answers)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': '分数已更新',
        'data': {
            'answer_id': answer.id,
            'score': float(answer.score) if answer.score else 0,
            'record_total_score': float(record.total_score) if record and record.total_score else 0
        }
    })


@admin_bp.route('/statistics/learning-data')
@login_or_jwt_required
@admin_required
def get_learning_statistics():
    progress_stats = db.session.query(
        CaseCategory.name,
        func.count(Station.id).label('total_stations'),
        func.count(LearningRecord.id).label('completed_stations'),
        func.avg(LearningRecord.score).label('avg_score')
    ).join(Case, CaseCategory.id == Case.category_id)\
     .join(Station, Case.id == Station.case_id)\
     .outerjoin(LearningRecord, Station.id == LearningRecord.station_id)\
     .group_by(CaseCategory.id, CaseCategory.name).all()

    progress_data = []
    for category_name, total, completed, avg_score in progress_stats:
        progress_data.append({
            'category': category_name,
            'total_stations': total,
            'completed_stations': completed or 0,
            'completion_rate': round((completed or 0) / total * 100, 1) if total > 0 else 0,
            'avg_score': round(float(avg_score), 1) if avg_score else 0
        })

    wrong_stats = db.session.query(
        CaseCategory.name,
        func.count(WrongQuestion.id).label('wrong_count')
    ).join(Case, CaseCategory.id == Case.category_id)\
     .join(Station, Case.id == Station.case_id)\
     .join(WrongQuestion, Station.id == WrongQuestion.station_id)\
     .group_by(CaseCategory.id, CaseCategory.name).all()

    wrong_data = {name: count for name, count in wrong_stats}

    user_activity = db.session.query(
        User.department,
        func.count(LearningRecord.id).label('activity_count'),
        func.avg(LearningRecord.score).label('avg_score')
    ).join(LearningRecord, User.id == LearningRecord.user_id)\
     .filter(User.role == 'nurse')\
     .group_by(User.department).all()

    activity_data = []
    for department, activity_count, avg_score in user_activity:
        activity_data.append({
            'department': department or '未分配科室',
            'activity_count': activity_count,
            'avg_score': round(float(avg_score), 1) if avg_score else 0
        })

    return jsonify({
        'success': True,
        'data': {
            'progress_stats': progress_data,
            'wrong_distribution': wrong_data,
            'user_activity': activity_data
        }
    })


@admin_bp.route('/statistics/group-weakness')
@login_or_jwt_required
@admin_required
def get_group_weakness():
    wrong_questions = db.session.query(WrongQuestion, Station, Case, CaseCategory)\
        .join(Station, WrongQuestion.station_id == Station.id)\
        .join(Case, Station.case_id == Case.id)\
        .join(CaseCategory, Case.category_id == CaseCategory.id).all()

    if not wrong_questions:
        return jsonify({
            'success': True,
            'data': {
                'analysis': {
                    'weak_categories': [],
                    'common_issues': ['暂无错题数据'],
                    'improvement_suggestions': [],
                    'priority_areas': []
                }
            }
        })

    category_errors = {}
    for wrong_q, station, case, category in wrong_questions:
        category_errors[category.name] = category_errors.get(category.name, 0) + 1

    weak_categories = sorted(category_errors.keys(), key=lambda x: category_errors[x], reverse=True)[:5]

    analysis = {
        'weak_categories': weak_categories,
        'common_issues': [f'{cat}领域错误率较高' for cat in weak_categories[:3]],
        'improvement_suggestions': [
            {'category': cat, 'suggestion': f'建议加强{cat}相关培训，重点关注常见错误'}
            for cat in weak_categories[:3]
        ],
        'priority_areas': weak_categories[:2],
        'error_distribution': category_errors
    }

    return jsonify({
        'success': True,
        'data': {
            'analysis': analysis,
            'total_errors': len(wrong_questions)
        }
    })


# ---- Station CRUD under Case ----

@admin_bp.route('/cases/<int:case_id>/stations', methods=['POST'])
@login_or_jwt_required
@admin_required
def create_case_station(case_id):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '站点名称不能为空'})
    max_order = db.session.query(func.max(Station.order_index)).filter_by(case_id=case_id).scalar() or 0
    station = Station(
        case_id=case_id, name=name,
        assessment_task=(data.get('assessment_task') or '').strip(),
        question=(data.get('question') or '').strip(),
        order_index=max_order + 1
    )
    db.session.add(station)
    db.session.commit()
    return jsonify({'success': True, 'station': {'id': station.id, 'name': station.name}})


@admin_bp.route('/cases/<int:case_id>/stations/<int:station_id>', methods=['PUT', 'DELETE'])
@login_or_jwt_required
@admin_required
def manage_case_station(case_id, station_id):
    station = Station.query.filter_by(id=station_id, case_id=case_id).first()
    if not station:
        return jsonify({'success': False, 'message': '站点不存在'}), 404

    if request.method == 'DELETE':
        try:
            # 删除关联记录（这些关系没有 cascade，需手动清除）
            from models import ExamAnswer
            WrongQuestion.query.filter_by(station_id=station_id).delete()
            LearningRecord.query.filter_by(station_id=station_id).delete()
            ExamAnswer.query.filter_by(station_id=station_id).delete()
            db.session.delete(station)
            db.session.commit()
            return jsonify({'success': True, 'message': '站点已删除'})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"站点删除失败: {e}", exc_info=True)
            return jsonify({'success': False, 'message': '删除失败，请稍后重试'}), 400

    data = request.get_json() or {}
    for field in ['name', 'assessment_task', 'question']:
        if field in data:
            setattr(station, field, (data.get(field) or '').strip())
    if 'order_index' in data:
        station.order_index = int(data['order_index'])
    db.session.commit()
    return jsonify({'success': True, 'message': '站点已更新'})


@admin_bp.route('/cases/<int:case_id>/stations/<int:station_id>/answers', methods=['PUT'])
@login_or_jwt_required
@admin_required
def update_station_answers(case_id, station_id):
    station = Station.query.filter_by(id=station_id, case_id=case_id).first()
    if not station:
        return jsonify({'success': False, 'message': '站点不存在'}), 404
    data = request.get_json() or {}
    items = data.get('answers') or []
    StandardAnswer.query.filter_by(station_id=station_id).delete()
    for i, item in enumerate(items):
        ans = StandardAnswer(
            station_id=station_id,
            answer_item=(item.get('answer_item') or '').strip(),
            score_weight=float(item.get('score_weight', 1.0)),
            order_index=i
        )
        db.session.add(ans)
    db.session.commit()
    return jsonify({'success': True, 'message': '答案已更新'})


# ---- Video Upload ----

@admin_bp.route('/videos/upload', methods=['POST'])
@login_or_jwt_required
@admin_required
@csrf.exempt
def upload_video_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})
    f = request.files['file']
    if f.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})
    import uuid
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.mp4', '.webm', '.ogg', '.mov', '.avi', '.mkv'):
        return jsonify({'success': False, 'message': f'不支持的视频格式: {ext}'})
    videos_dir = os.path.join(current_app.config.get('UPLOAD_DIR', '/app/uploads'), 'videos')
    os.makedirs(videos_dir, exist_ok=True)
    filename = f"{uuid.uuid4().hex}{ext}"
    f.save(os.path.join(videos_dir, filename))
    return jsonify({'success': True, 'url': f'/uploads/videos/{filename}'})


# ---- Video CRUD under Case ----

@admin_bp.route('/cases/<int:case_id>/videos', methods=['POST'])
@login_or_jwt_required
@admin_required
def create_case_video(case_id):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    url = (data.get('url') or '').strip()
    if not title or not url:
        return jsonify({'success': False, 'message': '视频标题和URL不能为空'})
    max_order = db.session.query(func.max(ExtensionVideo.order_index)).filter_by(case_id=case_id).scalar() or 0
    video = ExtensionVideo(case_id=case_id, title=title, url=url,
                           description=(data.get('description') or '').strip(),
                           order_index=max_order + 1)
    db.session.add(video)
    db.session.commit()
    return jsonify({'success': True, 'video': {'id': video.id, 'title': video.title}})


@admin_bp.route('/cases/<int:case_id>/videos/<int:video_id>', methods=['DELETE'])
@login_or_jwt_required
@admin_required
def delete_case_video(case_id, video_id):
    video = ExtensionVideo.query.filter_by(id=video_id, case_id=case_id).first()
    if not video:
        return jsonify({'success': False, 'message': '视频不存在'}), 404
    db.session.delete(video)
    db.session.commit()
    return jsonify({'success': True, 'message': '视频已删除'})


# ---- Link CRUD under Case ----

@admin_bp.route('/cases/<int:case_id>/links', methods=['POST'])
@login_or_jwt_required
@admin_required
def create_case_link(case_id):
    case = db.session.get(Case, case_id)
    if not case:
        return jsonify({'success': False, 'message': '案例不存在'}), 404
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    url = (data.get('url') or '').strip()
    if not title or not url:
        return jsonify({'success': False, 'message': '链接标题和URL不能为空'})
    max_order = db.session.query(func.max(ExtensionLink.order_index)).filter_by(case_id=case_id).scalar() or 0
    link = ExtensionLink(case_id=case_id, title=title, url=url,
                         description=(data.get('description') or '').strip(),
                         order_index=max_order + 1)
    db.session.add(link)
    db.session.commit()
    return jsonify({'success': True, 'link': {'id': link.id, 'title': link.title}})


@admin_bp.route('/cases/<int:case_id>/links/<int:link_id>', methods=['DELETE'])
@login_or_jwt_required
@admin_required
def delete_case_link(case_id, link_id):
    link = ExtensionLink.query.filter_by(id=link_id, case_id=case_id).first()
    if not link:
        return jsonify({'success': False, 'message': '链接不存在'}), 404
    db.session.delete(link)
    db.session.commit()
    return jsonify({'success': True, 'message': '链接已删除'})


# ---- Extended Knowledge (per-item) ----

@admin_bp.route('/cases/<int:case_id>/knowledge', methods=['POST'])
@login_or_jwt_required
@admin_required
def add_case_knowledge(case_id):
    Case.query.get_or_404(case_id)
    data = request.get_json()
    if not data or not data.get('question'):
        return jsonify({'success': False, 'message': '问题不能为空'}), 400
    answers = data.get('answers') or []
    if not answers:
        return jsonify({'success': False, 'message': '至少需要一个答案项'}), 400
    ek = ExtendedKnowledge(case_id=case_id, question=data['question'])
    db.session.add(ek)
    db.session.flush()
    for idx, a in enumerate(answers):
        db.session.add(KnowledgeAnswer(
            knowledge_id=ek.id,
            answer_item=(a.get('answer_item') or '').strip(),
            score_weight=float(a.get('score_weight', 1)),
            order_index=idx
        ))
    db.session.commit()
    return jsonify({'success': True, 'message': '扩展知识已添加'})


@admin_bp.route('/cases/<int:case_id>/knowledge/<int:knowledge_id>', methods=['DELETE'])
@login_or_jwt_required
@admin_required
def delete_case_knowledge(case_id, knowledge_id):
    ek = ExtendedKnowledge.query.filter_by(id=knowledge_id, case_id=case_id).first()
    if not ek:
        return jsonify({'success': False, 'message': '扩展知识不存在'}), 404
    db.session.delete(ek)
    db.session.commit()
    return jsonify({'success': True, 'message': '扩展知识已删除'})


# ---- User Progress ----

@admin_bp.route('/users/<int:user_id>/progress')
@login_or_jwt_required
@admin_required
def get_user_progress(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'}), 404

    # 按类别统计学习进度
    category_progress = db.session.query(
        CaseCategory.name,
        func.count(func.distinct(Station.id)).label('total'),
        func.count(func.distinct(LearningRecord.id)).label('completed')
    ).join(Case, CaseCategory.id == Case.category_id)\
     .join(Station, Case.id == Station.case_id)\
     .outerjoin(LearningRecord, (Station.id == LearningRecord.station_id) & (LearningRecord.user_id == user_id))\
     .filter(Case.case_type == 'learning')\
     .group_by(CaseCategory.id, CaseCategory.name).all()

    # 最近学习记录
    recent = db.session.query(LearningRecord, Station, Case)\
        .join(Station, LearningRecord.station_id == Station.id)\
        .join(Case, Station.case_id == Case.id)\
        .filter(LearningRecord.user_id == user_id)\
        .order_by(desc(LearningRecord.completed_at)).limit(20).all()

    # 考试记录
    exam_records = ExamRecord.query.filter_by(user_id=user_id)\
        .order_by(desc(ExamRecord.submit_time)).all()

    # 积分记录
    point_recs = PointRecord.query.filter_by(user_id=user_id)\
        .order_by(desc(PointRecord.created_at)).limit(20).all()

    return jsonify({'success': True, 'data': {
        'user': {
            'id': user.id, 'username': user.username, 'real_name': user.real_name,
            'department': user.department, 'points': user.points, 'status': user.status
        },
        'category_progress': [
            {'category': c, 'total': t, 'completed': co or 0,
             'progress': round((co or 0) / t * 100, 1) if t > 0 else 0}
            for c, t, co in category_progress
        ],
        'recent_records': [
            {'id': r.id, 'case_title': c.title, 'station_name': s.name,
             'score': float(r.score) if r.score else 0,
             'completed_at': r.completed_at.isoformat()}
            for r, s, c in recent
        ],
        'exam_records': [
            {'id': e.id, 'exam_id': e.exam_id, 'total_score': float(e.total_score) if e.total_score else 0,
             'status': e.status, 'submit_time': e.submit_time.isoformat() if e.submit_time else None}
            for e in exam_records
        ],
        'point_records': [
            {'id': p.id, 'points': p.points, 'reason': p.reason,
             'created_at': p.created_at.isoformat()}
            for p in point_recs
        ]
    }})


# ---- AI Test Connection ----

@admin_bp.route('/ai-settings/test', methods=['POST'])
@login_or_jwt_required
@admin_required
def test_ai_connection():
    import time
    data = request.get_json() or {}
    provider = data.get('provider', 'openai')
    api_key = data.get('api_key', '').strip()
    model = data.get('model', '').strip()
    base_url = data.get('base_url', '').strip()

    if not api_key:
        return jsonify({'success': False, 'message': '请提供 API Key'})

    try:
        start = time.time()
        if provider == 'openai':
            import openai
            openai.api_key = api_key
            if base_url:
                openai.api_base = base_url
            openai.ChatCompletion.create(
                model=model or 'gpt-4o-mini',
                messages=[{'role': 'user', 'content': 'ping'}],
                max_tokens=5
            )
        elif provider == 'glm':
            from zhipuai import ZhipuAI
            client_kwargs = {'api_key': api_key}
            if base_url:
                client_kwargs['base_url'] = base_url
            client = ZhipuAI(**client_kwargs)
            client.chat.completions.create(
                model=model or 'glm-4-air',
                messages=[{'role': 'user', 'content': 'ping'}],
                max_tokens=5
            )
        else:
            return jsonify({'success': False, 'message': '不支持的 provider'})
        latency = round((time.time() - start) * 1000)
        return jsonify({'success': True, 'message': f'连接成功，延迟 {latency}ms', 'latency_ms': latency})
    except Exception as e:
        current_app.logger.error(f"AI连接测试失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '连接失败，请稍后重试'})


# ---- Exam QR Code ----

@admin_bp.route('/exams/<int:exam_id>/qr-code')
@login_or_jwt_required
@admin_required
def get_exam_qr_code(exam_id):
    import qrcode
    from io import BytesIO
    from flask_jwt_extended import create_access_token

    exam = Exam.query.get_or_404(exam_id)
    token = create_access_token(identity=f'exam:{exam_id}')

    # 优先使用配置的 SITE_URL，否则从请求头探测
    site_url = current_app.config.get('SITE_URL', '')
    if site_url:
        base = site_url
    else:
        base = request.host_url.rstrip('/')
        # nginx 反向代理时，通过 X-Forwarded-Proto 还原协议
        proto = request.headers.get('X-Forwarded-Proto', '')
        if proto == 'https':
            base = base.replace('http://', 'https://')

    exam_url = f"{base}/nurse/exam-access?token={token}&exam_id={exam_id}"

    try:
        img = qrcode.make(exam_url)
        buf = BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        resp = send_file(buf, mimetype='image/png')
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp
    except Exception as e:
        logging.getLogger(__name__).error('QR 二维码生成失败：%s', e)
        return jsonify({'success': False, 'message': '二维码生成失败'}), 500


# ---- Exam Update ----

@admin_bp.route('/exams/<int:exam_id>', methods=['PUT'])
@login_or_jwt_required
@admin_required
def update_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    data = request.get_json() or {}
    for field in ['title', 'description']:
        if field in data:
            setattr(exam, field, (data.get(field) or '').strip())
    for field in ['duration']:
        if field in data:
            setattr(exam, field, int(data[field]))
    for field in ['start_time', 'end_time']:
        if field in data and data[field]:
            setattr(exam, field, datetime.fromisoformat(data[field]))
    db.session.commit()
    return jsonify({'success': True, 'message': '考试已更新'})


@admin_bp.route('/exams/<int:exam_id>/publish', methods=['POST'])
@login_or_jwt_required
@admin_required
def publish_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    exam.status = 'published'
    db.session.commit()
    return jsonify({'success': True, 'message': '考试已发布'})


# ---- Baidu ASR Key Management ----

@admin_bp.route('/baidu-asr-keys', methods=['GET'])
@login_or_jwt_required
@admin_required
def list_baidu_asr_keys():
    keys = BaiduAsrKey.query.order_by(BaiduAsrKey.id).all()
    return jsonify({'success': True, 'data': [{
        'id': k.id,
        'app_id': k.app_id,
        'api_key_masked': '***' + decrypt_value(k.api_key)[-4:] if decrypt_value(k.api_key) and len(decrypt_value(k.api_key)) > 4 else '***',
        'is_active': k.is_active,
        'created_at': k.created_at.isoformat(),
    } for k in keys]})


@admin_bp.route('/baidu-asr-keys', methods=['POST'])
@login_or_jwt_required
@admin_required
def add_baidu_asr_key():
    data = request.get_json() or {}
    app_id = data.get('app_id', '').strip()
    api_key = data.get('api_key', '').strip()
    secret_key = data.get('secret_key', '').strip()
    if not api_key or not secret_key:
        return jsonify({'success': False, 'message': 'API Key 和 Secret Key 不能为空'})
    try:
        k = BaiduAsrKey(
            app_id=app_id,
            api_key=encrypt_value(api_key),
            secret_key=encrypt_value(secret_key),
        )
        db.session.add(k)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Key 已添加', 'data': {'id': k.id}})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"百度ASR Key添加失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '添加失败，请稍后重试'})


@admin_bp.route('/baidu-asr-keys/<int:key_id>', methods=['DELETE'])
@login_or_jwt_required
@admin_required
def delete_baidu_asr_key(key_id):
    k = db.session.get(BaiduAsrKey, key_id)
    if not k:
        return jsonify({'success': False, 'message': 'Key 不存在'})
    try:
        db.session.delete(k)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Key 已删除'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"百度ASR Key删除失败: {e}", exc_info=True)
        return jsonify({'success': False, 'message': '删除失败，请稍后重试'})


@admin_bp.route('/baidu-asr-keys/<int:key_id>/toggle', methods=['POST'])
@login_or_jwt_required
@admin_required
def toggle_baidu_asr_key(key_id):
    k = db.session.get(BaiduAsrKey, key_id)
    if not k:
        return jsonify({'success': False, 'message': 'Key 不存在'})
    k.is_active = not k.is_active
    db.session.commit()
    return jsonify({'success': True, 'message': f'Key 已{"启用" if k.is_active else "禁用"}', 'is_active': k.is_active})
